import numpy as np
from requests import head
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import torch
import os
import pandas as pd
from openpyxl import load_workbook
import os.path
from os import path
from utils.model_utils import get_pretrained_model

from utils.constants import PATH_TO_RESULTS
# from code.utils.viz_utils import plot_test_predictions


def log_metrics(data, store, model, args, epoch, save_all=True):
    
    values = np.zeros([len(data[store]["test_data"])])
    preds = np.zeros([len(data[store]["test_data"])])

    for i in range(0, len(data[store]["test_dataset"]), args.pred_length):
        x, y = data[store]["test_dataset"][i]
        if args.model_type == "transformer":
            y_hat = model(x.unsqueeze(0), y.unsqueeze(0))
        else:
            y_hat = model(x.unsqueeze(0))
        
        if i == 0:
            x = data[store]["test_dataset"].scaler.inverse_transform(data[store]["test_dataset"][i][0]).squeeze()
            values[:args.history_length] = x
            preds[:args.history_length] = x
        
        y = data[store]["test_dataset"].scaler.inverse_transform(data[store]["test_dataset"][i][1]).squeeze()
        values[i+args.history_length:i+args.history_length+args.pred_length] = y
        if args.model_type == "lstm" or args.model_type == "tcn":
            y_hat = data[store]["test_dataset"].scaler.inverse_transform(y_hat.detach().numpy()).squeeze()
        elif args.model_type == "dense" or args.model_type == "transformer":
            y_hat = data[store]["test_dataset"].scaler.inverse_transform(y_hat.detach().numpy().reshape([-1, 1])).squeeze()
        y_hat = np.round(y_hat)
        preds[i+args.history_length:i+args.history_length+args.pred_length] = y_hat

    mae = mean_absolute_error(values, preds)
    mse = mean_squared_error(values, preds)
    rmse = np.sqrt(mse)
    r2 = r2_score(values, preds)
    
    if not save_all:
        return np.around(mae, decimals=5), np.around(mse, decimals=5), np.around(rmse, decimals=5), np.around(r2, decimals=5)
    
    string = "Epoch: {}\nMAE: {}\nMSE: {}\nRMSE: {}\nR2: {}\n".format(epoch, mae, mse, rmse, r2)
    with open(args.path_to_logs[store] + '/metrics.txt', 'w') as f:
        f.write(string)
        
        
def eval_epoch(model, data, store, args, epoch, criterion, best_loss):
    model.eval()
    test_epoch_loss = np.zeros(len(data[store]["test_loader"]))
    test_epoch_loss_rescaled = np.zeros(len(data[store]["test_loader"]))
    with torch.no_grad():
        for i, (x, y) in enumerate(data[store]["test_loader"]):
            # loss
            if args.model_type == "transformer":
                y_hat = model(x, y)
            else:
                y_hat = model(x)
            
            loss = criterion(y_hat, y.squeeze())
        
            test_epoch_loss[i] = loss.item()
            
            # rescaled input
            y_hat = torch.tensor(data[store]["train_dataset"].scaler.inverse_transform(y_hat))
            y_hat = torch.round(y_hat)
            
            y = torch.tensor(data[store]["train_dataset"].scaler.inverse_transform(y.squeeze()))
            
            act_loss = criterion(y_hat, y)
            test_epoch_loss_rescaled[i] = act_loss.item()
            
    test_mean_loss = np.mean(test_epoch_loss)
    test_mean_loss_rescaled = np.mean(test_epoch_loss_rescaled)
    #Early stopping with log
    if test_mean_loss_rescaled < best_loss:
        print("best loss achived! Saving model...\n old: {}, new: {}\n".format(best_loss, test_mean_loss_rescaled))
        best_loss = test_mean_loss_rescaled
        torch.save(model.state_dict(), args.path_to_ckpt[store])
        string = 'Epoch: {}, Test MSE: {}'.format(epoch+1, test_mean_loss_rescaled)
        with open(os.path.join(args.path_to_logs[store], 'ckpt.txt'), 'w') as f:
            f.write(string)
        log_metrics(data, store, model, args, epoch+1)
        #plot_test_predictions(data, store, model, args, "best")
    
    string = 'Epoch [{}/{}], Loss: {:.4f}, Actual distance {}'.format(epoch+1, args.epochs, test_mean_loss, test_mean_loss_rescaled)
    print(string)
    with open(os.path.join(args.path_to_logs[store], 'test.txt'), 'a') as f:
        f.write(string + '\n')
        
    return best_loss


def save_results(data, store, args, epoch):
    model = get_pretrained_model(store=store)
    MAE, MSE, RMSE, R2 = log_metrics(data, store, model, args, epoch, save_all=False)
    data = [[args.model_type, args.name, args.epochs, args.history_length, args.pred_length, MAE, MSE, RMSE, R2]]
    df = pd.DataFrame(data, columns=["model_type", "name", "epochs", "history_length", "pred_length", "MAE", "MSE", "RMSE", "R2"])
    if path.exists(PATH_TO_RESULTS):
        FilePath = PATH_TO_RESULTS
        ExcelWorkbook = load_workbook(FilePath)
        writer = pd.ExcelWriter(FilePath, engine = 'openpyxl')
        writer.book = ExcelWorkbook
        writer.sheets =  {ws.title: ws for ws in writer.book.worksheets}
        if store not in writer.sheets:
            ExcelWorkbook.create_sheet(title=store)
            writer.sheets =  {ws.title: ws for ws in writer.book.worksheets}
            # startrow = writer.sheets[store].max_row
            df.to_excel(writer, sheet_name=store, index = False, header= True)
        else:
            startrow = writer.sheets[store].max_row
            df.to_excel(writer, sheet_name=store, startrow = startrow, index = False, header= False)
        # startr,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
        writer.save()
        writer.close()
    else:
        df.to_excel(PATH_TO_RESULTS, sheet_name=store, index=False, header=True)